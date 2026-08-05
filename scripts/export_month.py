"""Export one month's dashboard-data JSON from an FP&A workbook.

Usage:
    python scripts/export_month.py "workbook.xlsx"
    python scripts/export_month.py "workbook.xlsx" --monthly-detail "monthly_opex.xlsx"
    python scripts/export_month.py "workbook.xlsx" --refresh-config

Primary source:
  Workbook's `Dashboard JSON` sheet — cells A2 (+A3 if present) hold the already-
  assembled dashboard-data.json (Excel splits it across cells because of the
  32k-char cell limit).

Layout config (--refresh-config only):
  Workbook's `Dashboard Config` A2. Off by default so hand edits to the layout
  survive re-running the export.

Software vendor detail:
  Scans each `BvA <Leader>` tab for the 'Software BvA Summary' section and
  attaches its vendors + monthly detail to the corresponding leader record.

Finance commentary (--monthly-detail only):
  Reads the separate 'Opex Monthly Summary' workbook (a QBO-derived pivot).
  Pulls the human-written 'Accounting Notes' per Class (=department), maps
  Class codes via scripts/class-to-department.json, and computes top MoM
  drivers per department. Attaches:
      commentary.raw     — list of finance-authored notes with amounts
      commentary.movers  — top MoM $ deltas per department
      commentary.summary — empty; filled in by editorial pass
  Also aggregates children's notes up onto each leader.

Writes:
    sample-data/dashboard-data-<YYYY-MM>.json          — that month's numbers
    sample-data/dashboard-data-index.json              — list of months + default
    sample-data/dashboard-config.json                  — latest layout

The month key (YYYY-MM) is derived from meta.reportMonth in the JSON so file
names match the numbers regardless of what the workbook file is named.

Requires openpyxl (`pip install --user openpyxl`).
"""
import json
import re
import shutil
import sys
import tempfile
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed. Run: pip install --user openpyxl")


REPO_ROOT = Path(__file__).resolve().parent.parent
SAMPLE_DIR = REPO_ROOT / "sample-data"

MONTH_NAMES = {
    "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12,
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "jun": 6, "jul": 7, "aug": 8,
    "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dec": 12,
}


def month_key_from_report_month(report_month: str) -> str:
    """Turn 'June 2026' or '2026-06' etc. into '2026-06'."""
    rm = str(report_month).strip()
    m = re.match(r"^(\d{4})[-/](\d{1,2})", rm)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}"
    m = re.match(r"^([A-Za-z]+)\s+(\d{4})$", rm)
    if m:
        month = MONTH_NAMES.get(m.group(1).lower())
        if month:
            return f"{m.group(2)}-{month:02d}"
    raise ValueError(f"Could not parse reportMonth: {report_month!r}")


def read_sheet_cell_concat(wb, sheet_name: str, col: int = 1) -> str:
    """Concatenate every non-empty cell in `col` after row 1 (row 1 is the
    'this is the JSON blob' comment). Excel splits long values across rows
    because of the 32,767 char per-cell limit."""
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet {sheet_name!r} not found in workbook")
    ws = wb[sheet_name]
    parts = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v is not None:
            parts.append(str(v))
    if not parts:
        raise ValueError(f"{sheet_name!r} col {col} row 2+ is empty")
    return "".join(parts)


# The Software BvA Summary section lives at slightly different rows in every
# leader's BvA tab. Layout (constant across leaders, verified against workbook):
#   label row:            'Software BvA Summary' text
#   label+1 (period row): 'Current Month <Mon>-<YY> v Budget', 'QTD ...', 'YTD ...'
#   label+2 (header row): Vendor | Department | Actual | Budget | ... (repeats for each period)
#
# The period row lets us disambiguate the three 'Actual'/'Budget' pairs that
# would otherwise collide in a naive column-name lookup.


def _num(v):
    return float(v) if isinstance(v, (int, float)) else None


def _snapshot_rows(ws, up_to_row: int = 260) -> list[list]:
    """Materialize a rectangle of the sheet into memory once. openpyxl's
    read_only mode makes iter_rows fast (streaming) but ws.cell(r,c) slow —
    each call restarts iteration from the top. Cache once, then random-access."""
    rows: list[list] = []
    for row in ws.iter_rows(min_row=1, max_row=up_to_row, values_only=True):
        rows.append(list(row))
    return rows


def _find_first_label_row(snap: list[list], label: str, max_col: int = 12) -> int | None:
    """1-based row index of the first row whose col-1..max_col contain `label`."""
    for r, row in enumerate(snap, start=1):
        for v in row[:max_col]:
            if isinstance(v, str) and v.strip() == label:
                return r
    return None


def _period_columns(snap: list[list], label_row: int) -> dict[str, int]:
    """Locate 1-based start column of each period block. Row(s) below the
    'Software BvA Summary' label carry period headers like 'Current Month
    <Mon>-<YY> v Budget', 'QTD ...', 'YTD ...'. Each block's Actual sits at the
    period column and Budget one column over."""
    ranges: dict[str, int] = {}
    for r in range(label_row, min(label_row + 4, len(snap) + 1)):
        row = snap[r - 1]
        for c_idx, v in enumerate(row, start=1):
            if not isinstance(v, str):
                continue
            s = v.strip().lower()
            if "current month" in s and "mtd" not in ranges:
                ranges["mtd"] = c_idx
            elif s.startswith("qtd") and "qtd" not in ranges:
                ranges["qtd"] = c_idx
            elif s.startswith("ytd") and "ytd" not in ranges:
                ranges["ytd"] = c_idx
        if ranges:
            break
    return ranges


def _cell(snap: list[list], r: int, c: int):
    """1-based access into the row cache. Returns None past end of row/sheet."""
    if r < 1 or r > len(snap):
        return None
    row = snap[r - 1]
    if c < 1 or c > len(row):
        return None
    return row[c - 1]


def _window_from_cols(snap, r: int, actual_col: int, budget_col: int) -> dict:
    a = _num(_cell(snap, r, actual_col))
    b = _num(_cell(snap, r, budget_col))
    var_pct = ((a - b) / b) if (a is not None and b not in (None, 0)) else None
    return {"actual": a, "budget": b, "varPct": var_pct}


def _extract_monthly_series(snap: list[list], label: str) -> dict[str, list[dict]]:
    """Read a 12-month table ('Monthly Actuals Software Summary' or
    'Monthly Budget Software Summary'). Returns {vendor_name: [{month, value}]}.
    Month strings are YYYY-MM."""
    label_row = _find_first_label_row(snap, label)
    if label_row is None:
        return {}
    header_row = None
    vendor_col = None
    date_cols: list[tuple[int, str]] = []
    for r in range(label_row, min(label_row + 6, len(snap) + 1)):
        row = snap[r - 1]
        if not any(isinstance(v, str) and v.strip() == "Vendor" for v in row):
            continue
        header_row = r
        for c_idx, v in enumerate(row, start=1):
            if isinstance(v, str) and v.strip() == "Vendor":
                vendor_col = c_idx
            if hasattr(v, "year"):
                date_cols.append((c_idx, f"{v.year:04d}-{v.month:02d}"))
        break
    if header_row is None or vendor_col is None or not date_cols:
        return {}

    out: dict[str, list[dict]] = {}
    for r in range(header_row + 1, min(header_row + 100, len(snap) + 1)):
        name = _cell(snap, r, vendor_col)
        if name is None or (isinstance(name, str) and not name.strip()):
            break
        name_s = str(name).strip()
        out[name_s] = [
            {"month": m, "value": _num(_cell(snap, r, c))} for c, m in date_cols
        ]
        if name_s == "Total":
            break
    return out


def extract_software_vendors(ws) -> list[dict] | None:
    # Snapshot the top of the sheet in one pass. Every subsequent lookup is a
    # cheap Python list indexing operation.
    snap = _snapshot_rows(ws, up_to_row=260)

    label_row = _find_first_label_row(snap, "Software BvA Summary")
    if label_row is None:
        return None

    periods = _period_columns(snap, label_row)
    if "mtd" not in periods:
        return None

    header_row = None
    vendor_col = dept_col = lastmo_col = None
    for r in range(label_row, min(label_row + 6, len(snap) + 1)):
        row = snap[r - 1]
        if not any(isinstance(v, str) and v.strip() == "Vendor" for v in row):
            continue
        header_row = r
        for c_idx, v in enumerate(row, start=1):
            if not isinstance(v, str):
                continue
            key = v.strip()
            if key == "Vendor" and vendor_col is None:
                vendor_col = c_idx
            elif key == "Department" and dept_col is None:
                dept_col = c_idx
            elif key == "LastMo Actual" and lastmo_col is None:
                lastmo_col = c_idx
        break
    if header_row is None or vendor_col is None:
        return None

    monthly_actual = _extract_monthly_series(snap, "Monthly Actuals Software Summary")
    monthly_budget = _extract_monthly_series(snap, "Monthly Budget Software Summary")

    vendors: list[dict] = []
    for r in range(header_row + 1, min(header_row + 100, len(snap) + 1)):
        name = _cell(snap, r, vendor_col)
        if name is None or (isinstance(name, str) and not name.strip()):
            break
        name_s = str(name).strip()

        mtd = _window_from_cols(snap, r, periods["mtd"], periods["mtd"] + 1)
        qtd = _window_from_cols(snap, r, periods["qtd"], periods["qtd"] + 1) if "qtd" in periods else None
        ytd = _window_from_cols(snap, r, periods["ytd"], periods["ytd"] + 1) if "ytd" in periods else None

        if (
            mtd["actual"] is None
            and mtd["budget"] is None
            and name_s not in {"Total", "All Other"}
        ):
            continue

        lm = _num(_cell(snap, r, lastmo_col)) if lastmo_col else None
        mom_delta = (mtd["actual"] - lm) if (mtd["actual"] is not None and lm is not None) else None
        mom_pct = (mom_delta / lm) if (mom_delta is not None and lm not in (None, 0)) else None

        dept = _cell(snap, r, dept_col) if dept_col else None
        row_out = {
            "name": name_s,
            "department": str(dept).strip() if isinstance(dept, str) else None,
            "isTotal": name_s == "Total",
            "isOther": name_s == "All Other",
            "mtd": mtd,
            "qtd": qtd,
            "ytd": ytd,
            "lastMonthActual": lm,
            "mom": {"delta": mom_delta, "pct": mom_pct},
            "monthlyActual": monthly_actual.get(name_s),
            "monthlyBudget": monthly_budget.get(name_s),
        }
        vendors.append(row_out)
        if name_s == "Total":
            break

    return vendors or None


# =====================================================================
# Commentary extraction from the monthly Opex file (QBO pivot)
# =====================================================================


def _load_class_to_dept(path: Path) -> dict[str, str]:
    """Class label (as written in the QBO file) → department dataKey."""
    raw = json.loads(path.read_text(encoding="utf-8"))
    return {k: v for k, v in raw.items() if not k.startswith("_")}


def _find_header_row(snap: list[list], must_have: list[str]) -> int | None:
    """First 1-based row where every string in `must_have` appears as a cell."""
    needed = set(must_have)
    for r_idx, row in enumerate(snap, start=1):
        seen = {str(v).strip() for v in row if isinstance(v, str)}
        if needed.issubset(seen):
            return r_idx
    return None


def _snapshot_full_sheet(ws, max_rows: int = 400) -> list[list]:
    rows: list[list] = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_rows), values_only=True):
        rows.append(list(row))
    return rows


def _extract_notes_from_pivot(
    ws,
    class_map: dict[str, str],
    max_rows: int = 400,
) -> tuple[dict[str, list[dict]], list[str]]:
    """Return {department: [{class, account, mtdActual, variance, note}]} and
    a list of Class labels seen but not mapped. Reads a pivot with headers
    Class + Account + monthly date columns + Variance + a notes column."""
    snap = _snapshot_full_sheet(ws, max_rows)

    hdr_row = _find_header_row(snap, ["Class", "Account"])
    if hdr_row is None:
        return {}, []

    hdr = snap[hdr_row - 1]
    col_idx = {v.strip(): i + 1 for i, v in enumerate(hdr) if isinstance(v, str)}
    class_col = col_idx.get("Class")
    account_col = col_idx.get("Account")
    if not (class_col and account_col):
        return {}, []

    # Notes columns: look for "Accounting Notes" / "Notes" / any col with
    # "note" in its header. Skip "Y/N" flag columns.
    notes_col = None
    variance_col = None
    for label, ci in col_idx.items():
        lo = label.lower()
        if notes_col is None and ("note" in lo or "commentary" in lo) and "y/n" not in lo:
            notes_col = ci
        if variance_col is None and "variance" in lo and "accounting variance" not in lo:
            variance_col = ci
    # Fallback: 'Accounting Variance' is the notes column in EXPENSES BY MONTH SUMMARY
    if notes_col is None:
        for label, ci in col_idx.items():
            if label.strip().lower() == "accounting variance":
                notes_col = ci
                break

    # Month columns are datetimes in the header row; the LAST one is the
    # current reporting month, second-to-last is prior. Use them to compute
    # MoM $ delta for each row.
    month_cols: list[tuple[int, str]] = []
    for i, v in enumerate(hdr, start=1):
        if hasattr(v, "year"):
            month_cols.append((i, f"{v.year:04d}-{v.month:02d}"))
    curr_col = month_cols[-1][0] if month_cols else None
    prev_col = month_cols[-2][0] if len(month_cols) >= 2 else None

    by_dept: dict[str, list[dict]] = {}
    unmapped: set[str] = set()

    for r in range(hdr_row + 1, len(snap) + 1):
        class_val = snap[r - 1][class_col - 1] if class_col - 1 < len(snap[r - 1]) else None
        if not isinstance(class_val, str) or not class_val.strip():
            continue
        cls = class_val.strip()
        if cls in {"Grand Total"} or cls.endswith(" Total"):
            continue

        dept = class_map.get(cls)
        note = snap[r - 1][notes_col - 1] if (notes_col and notes_col - 1 < len(snap[r - 1])) else None
        if not (isinstance(note, str) and note.strip()):
            # Skip rows that don't carry finance commentary; the raw amounts
            # are already captured elsewhere in the JSON.
            continue

        if dept is None:
            unmapped.add(cls)
            continue

        account = snap[r - 1][account_col - 1] if account_col - 1 < len(snap[r - 1]) else None
        mtd_actual = None
        prior_actual = None
        if curr_col and curr_col - 1 < len(snap[r - 1]):
            v = snap[r - 1][curr_col - 1]
            mtd_actual = float(v) if isinstance(v, (int, float)) else None
        if prev_col and prev_col - 1 < len(snap[r - 1]):
            v = snap[r - 1][prev_col - 1]
            prior_actual = float(v) if isinstance(v, (int, float)) else None
        mom_delta = (
            mtd_actual - prior_actual
            if (mtd_actual is not None and prior_actual is not None)
            else None
        )

        by_dept.setdefault(dept, []).append({
            "class": cls,
            "account": str(account).strip() if isinstance(account, str) else None,
            "mtdActual": mtd_actual,
            "priorActual": prior_actual,
            "momDelta": mom_delta,
            "note": note.strip(),
        })

    return by_dept, sorted(unmapped)


def _compute_movers_by_dept(
    ws, class_map: dict[str, str], top_n: int = 5, max_rows: int = 400,
) -> dict[str, list[dict]]:
    """Regardless of notes, pull the top ±MoM movers by $ per department.
    Uses the same Class + Account + monthly-date pivot layout."""
    snap = _snapshot_full_sheet(ws, max_rows)
    hdr_row = _find_header_row(snap, ["Class", "Account"])
    if hdr_row is None:
        return {}
    hdr = snap[hdr_row - 1]
    col_idx = {v.strip(): i + 1 for i, v in enumerate(hdr) if isinstance(v, str)}
    class_col = col_idx.get("Class")
    account_col = col_idx.get("Account")
    month_cols = [(i, f"{v.year:04d}-{v.month:02d}") for i, v in enumerate(hdr, start=1) if hasattr(v, "year")]
    if not (class_col and account_col and len(month_cols) >= 2):
        return {}
    curr_col = month_cols[-1][0]
    prev_col = month_cols[-2][0]

    by_dept: dict[str, list[dict]] = {}
    for r in range(hdr_row + 1, len(snap) + 1):
        class_val = snap[r - 1][class_col - 1] if class_col - 1 < len(snap[r - 1]) else None
        if not isinstance(class_val, str):
            continue
        cls = class_val.strip()
        if cls.endswith(" Total") or cls in {"Grand Total"}:
            continue
        dept = class_map.get(cls)
        if dept is None:
            continue
        curr = snap[r - 1][curr_col - 1] if curr_col - 1 < len(snap[r - 1]) else None
        prev = snap[r - 1][prev_col - 1] if prev_col - 1 < len(snap[r - 1]) else None
        curr_f = float(curr) if isinstance(curr, (int, float)) else None
        prev_f = float(prev) if isinstance(prev, (int, float)) else None
        if curr_f is None or prev_f is None:
            continue
        delta = curr_f - prev_f
        if abs(delta) < 500:  # noise floor
            continue
        account = snap[r - 1][account_col - 1] if account_col - 1 < len(snap[r - 1]) else None
        by_dept.setdefault(dept, []).append({
            "account": str(account).strip() if isinstance(account, str) else None,
            "mtdActual": curr_f,
            "priorActual": prev_f,
            "momDelta": delta,
            "momPct": (delta / prev_f) if prev_f else None,
        })
    # Keep only the top |delta| N per department.
    for dept, rows in by_dept.items():
        rows.sort(key=lambda x: abs(x["momDelta"] or 0), reverse=True)
        by_dept[dept] = rows[:top_n]
    return by_dept


def attach_commentary(data: dict, monthly_path: Path, class_map_path: Path) -> None:
    """Merge finance notes + top movers from the monthly file into each
    department/leader record in `data` (in place)."""
    class_map = _load_class_to_dept(class_map_path)
    wb = load_workbook_unlocked(monthly_path)

    # Notes: pull from the two "*Expenses by Account" pivots and the roll-up
    # sheet. Merge across sheets; dedupe by (class, account, note).
    seen: set[tuple] = set()
    combined_notes: dict[str, list[dict]] = {}
    all_unmapped: set[str] = set()
    for sheet_name in [
        "Opex Expenses by Account ",
        "COGS Expenses by Account",
        "EXPENSES BY MONTH SUMMARY",
    ]:
        if sheet_name not in wb.sheetnames:
            continue
        notes, unmapped = _extract_notes_from_pivot(wb[sheet_name], class_map)
        for dept, rows in notes.items():
            for row in rows:
                key = (dept, row.get("class"), row.get("account"), row.get("note"))
                if key in seen:
                    continue
                seen.add(key)
                combined_notes.setdefault(dept, []).append(row)
        all_unmapped.update(unmapped)

    if all_unmapped:
        print(
            "  commentary: unmapped Class labels "
            "(add to scripts/class-to-department.json): "
            + ", ".join(sorted(all_unmapped))
        )

    # Movers: pull from the Opex + COGS pivots. Merge, then trim to top-5.
    all_movers: dict[str, list[dict]] = {}
    for sheet_name in ["Opex Expenses by Account ", "COGS Expenses by Account"]:
        if sheet_name not in wb.sheetnames:
            continue
        mv = _compute_movers_by_dept(wb[sheet_name], class_map)
        for dept, rows in mv.items():
            all_movers.setdefault(dept, []).extend(rows)
    for dept in list(all_movers.keys()):
        all_movers[dept].sort(key=lambda x: abs(x["momDelta"] or 0), reverse=True)
        all_movers[dept] = all_movers[dept][:5]

    # Attach to each department record. Missing = empty block so the frontend
    # can still render the section without checking for presence.
    for dept in data.get("departments", []):
        name = dept["name"]
        dept["commentary"] = {
            "raw": combined_notes.get(name, []),
            "movers": all_movers.get(name, []),
            "summary": [],
        }

    # For leaders, roll up their children's commentary (from dashboard-config
    # `children` — same source truth used by breakdownTable). Load the config
    # once; if it's not present (or a leader lacks a `children` array), skip.
    cfg_path = SAMPLE_DIR / "dashboard-config.json"
    child_by_leader: dict[str, list[str]] = {}
    if cfg_path.exists():
        cfg = json.loads(cfg_path.read_text(encoding="utf-8"))
        for entry in cfg.get("dashboards", []):
            if entry.get("source") == "leaders" and entry.get("children"):
                child_by_leader[entry["dataKey"]] = list(entry["children"])

    for leader in data.get("leaders", []):
        children = child_by_leader.get(leader["name"], [])
        combined_raw: list[dict] = []
        combined_mov: list[dict] = []
        for child in children:
            for row in combined_notes.get(child, []):
                combined_raw.append({**row, "department": child})
            for row in all_movers.get(child, []):
                combined_mov.append({**row, "department": child})
        # Order movers by absolute MoM $ delta.
        combined_mov.sort(key=lambda x: abs(x.get("momDelta") or 0), reverse=True)
        leader["commentary"] = {
            "raw": combined_raw,
            "movers": combined_mov[:8],  # a leader's rollup can afford a few more
            "summary": [],
            "children": children,
        }

    total_notes = sum(len(v) for v in combined_notes.values())
    print(f"  commentary: {total_notes} finance notes across {len(combined_notes)} depts; "
          f"{sum(len(v) for v in all_movers.values())} movers")


def load_workbook_unlocked(xlsx_path: Path):
    """openpyxl can't open files locked by Excel (e.g., open in OneDrive). If the
    direct open fails with PermissionError, copy to a temp file and retry."""
    try:
        return openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    except PermissionError:
        tmp = Path(tempfile.mkdtemp()) / xlsx_path.name
        shutil.copy2(xlsx_path, tmp)
        return openpyxl.load_workbook(tmp, data_only=True, read_only=True)


def main() -> int:
    argv = sys.argv[1:]
    args: list[str] = []
    monthly_detail: str | None = None
    flags: set[str] = set()
    i = 0
    while i < len(argv):
        a = argv[i]
        if a == "--monthly-detail":
            if i + 1 >= len(argv):
                sys.exit("--monthly-detail requires a path argument")
            monthly_detail = argv[i + 1]
            i += 2
            continue
        if a.startswith("--"):
            flags.add(a)
            i += 1
            continue
        args.append(a)
        i += 1

    if len(args) != 1:
        sys.exit(
            "Usage: python scripts/export_month.py <workbook.xlsx> "
            "[--monthly-detail <opex_monthly.xlsx>] [--refresh-config]"
        )
    refresh_config = "--refresh-config" in flags
    src = Path(args[0]).expanduser().resolve()
    if not src.exists():
        sys.exit(f"File not found: {src}")

    wb = load_workbook_unlocked(src)

    data_raw = read_sheet_cell_concat(wb, "Dashboard JSON")
    data = json.loads(data_raw)

    if "meta" not in data or "reportMonth" not in data["meta"]:
        sys.exit("Dashboard JSON is missing meta.reportMonth")

    month_key = month_key_from_report_month(data["meta"]["reportMonth"])

    # Attach Software BvA Summary vendor rows to each leader. Each leader's tab
    # is named 'BvA <Name>' — same convention the workbook uses. Missing sheets
    # or missing sections are logged, not fatal.
    for leader in data.get("leaders", []):
        name = leader.get("name")
        # sourceTab exists in the JSON already (e.g. "BvA Armaan"); fall back to
        # the standard convention if not.
        sheet_name = leader.get("sourceTab") or f"BvA {name}"
        if sheet_name not in wb.sheetnames:
            print(f"  software: skipped {name} (no sheet {sheet_name!r})")
            continue
        vendors = extract_software_vendors(wb[sheet_name])
        if not vendors:
            print(f"  software: no 'Software BvA Summary' found in {sheet_name}")
            continue
        leader["software"] = {"vendors": vendors}
        # Count excludes the Total row for the log line
        real_vendors = [v for v in vendors if not v.get("isTotal")]
        print(f"  software: {name} — {len(real_vendors)} vendors + total")

    # Optional: pull finance-authored notes + top MoM movers from the separate
    # monthly detail file (QBO-derived pivot). Only runs when --monthly-detail
    # is passed; departments and leaders still export cleanly without it.
    if monthly_detail:
        detail_path = Path(monthly_detail).expanduser().resolve()
        if not detail_path.exists():
            sys.exit(f"--monthly-detail file not found: {detail_path}")
        class_map_path = Path(__file__).resolve().parent / "class-to-department.json"
        if not class_map_path.exists():
            sys.exit(f"Missing class map: {class_map_path}")
        attach_commentary(data, detail_path, class_map_path)

    SAMPLE_DIR.mkdir(parents=True, exist_ok=True)
    data_path = SAMPLE_DIR / f"dashboard-data-{month_key}.json"
    data_path.write_text(json.dumps(data, indent=2), encoding="utf-8")
    print(f"wrote {data_path.relative_to(REPO_ROOT)} "
          f"({len(data.get('departments', []))} depts, "
          f"{len(data.get('leaders', []))} leaders)")

    # Layout config: only refresh from the workbook on explicit request. Otherwise
    # leave the checked-in dashboard-config.json alone so hand-edited templates,
    # children maps, and dashboard lists survive a data re-export.
    if refresh_config:
        try:
            cfg_raw = read_sheet_cell_concat(wb, "Dashboard Config")
            cfg = json.loads(cfg_raw)
            cfg_path = SAMPLE_DIR / "dashboard-config.json"
            cfg_path.write_text(json.dumps(cfg, indent=2), encoding="utf-8")
            print(f"wrote {cfg_path.relative_to(REPO_ROOT)} "
                  f"({len(cfg.get('dashboards', []))} dashboards)")
        except (KeyError, ValueError) as e:
            print(f"skipped config (workbook may not carry it): {e}")
    else:
        print("config: skipped (pass --refresh-config to overwrite from workbook)")

    # Rebuild the index from every dashboard-data-*.json on disk. Newest first.
    index_path = SAMPLE_DIR / "dashboard-data-index.json"
    months = []
    for p in sorted(SAMPLE_DIR.glob("dashboard-data-*.json"), reverse=True):
        if p.name == "dashboard-data-index.json":
            continue
        key = p.stem.removeprefix("dashboard-data-")
        try:
            payload = json.loads(p.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            continue
        months.append({
            "key": key,
            "label": payload.get("meta", {}).get("reportMonth", key),
            "source": payload.get("meta", {}).get("source"),
        })

    index = {
        "months": months,
        "default": months[0]["key"] if months else None,
    }
    index_path.write_text(json.dumps(index, indent=2), encoding="utf-8")
    print(f"wrote {index_path.relative_to(REPO_ROOT)} "
          f"(months: {', '.join(m['key'] for m in months)}; default {index['default']})")

    return 0


if __name__ == "__main__":
    sys.exit(main())
